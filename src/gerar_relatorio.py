import pandas as pd
import openpyxl
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill


def gerar_relatorio(caminho_csv="output/base_satisfacao_limpa.csv", caminho_saida="output/analise_satisfacao.xlsx"):
    df = pd.read_csv(caminho_csv)

    por_trimestre = df.groupby("trimestre")["nota_satisfacao"].mean().round(2).reset_index()
    por_produto = (
        df.groupby(["trimestre", "produto"])["nota_satisfacao"]
        .mean()
        .round(2)
        .unstack()
        .reset_index()
    )

    wb = openpyxl.Workbook()

    titulo_font = Font(name="Arial", size=13, bold=True)
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    body_font = Font(name="Arial", size=11)

    # ===== Aba 1: nota media por trimestre =====
    ws1 = wb.active
    ws1.title = "Nota por Trimestre"
    ws1["A1"] = "Nota média de satisfação por trimestre"
    ws1["A1"].font = titulo_font

    ws1["A3"] = "Trimestre"
    ws1["B3"] = "Nota média"
    for cell in ["A3", "B3"]:
        ws1[cell].font = header_font
        ws1[cell].fill = header_fill
        ws1[cell].alignment = Alignment(horizontal="center")

    for i, row in por_trimestre.iterrows():
        ws1.cell(row=4 + i, column=1, value=str(row["trimestre"])).font = body_font
        ws1.cell(row=4 + i, column=2, value=float(row["nota_satisfacao"])).font = body_font

    ws1.column_dimensions["A"].width = 14
    ws1.column_dimensions["B"].width = 12

    chart1 = LineChart()
    chart1.title = "Evolução da nota média de satisfação"
    chart1.style = 12
    chart1.y_axis.title = "Nota média (0-10)"
    chart1.x_axis.title = "Trimestre"
    chart1.height = 9
    chart1.width = 18

    data = Reference(ws1, min_col=2, min_row=3, max_row=3 + len(por_trimestre))
    cats = Reference(ws1, min_col=1, min_row=4, max_row=3 + len(por_trimestre))
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.series[0].marker.symbol = "circle"

    ws1.add_chart(chart1, "D3")

    # ===== Aba 2: nota media por produto e trimestre =====
    ws2 = wb.create_sheet("Nota por Produto")
    ws2["A1"] = "Nota média de satisfação por produto e trimestre"
    ws2["A1"].font = titulo_font

    colunas = list(por_produto.columns)
    for j, col in enumerate(colunas):
        cell = ws2.cell(row=3, column=1 + j, value=str(col))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, row in por_produto.iterrows():
        for j, col in enumerate(colunas):
            val = str(row[col]) if col == "trimestre" else float(row[col])
            ws2.cell(row=4 + i, column=1 + j, value=val).font = body_font

    for col_letter, width in zip("ABCDEF", [12, 10, 10, 22, 10, 16]):
        ws2.column_dimensions[col_letter].width = width

    chart2 = BarChart()
    chart2.type = "col"
    chart2.grouping = "clustered"
    chart2.title = "Nota média por produto, ao longo dos trimestres"
    chart2.style = 10
    chart2.y_axis.title = "Nota média (0-10)"
    chart2.x_axis.title = "Trimestre"
    chart2.height = 10
    chart2.width = 20

    data2 = Reference(ws2, min_col=2, max_col=1 + len(colunas) - 1, min_row=3, max_row=3 + len(por_produto))
    cats2 = Reference(ws2, min_col=1, min_row=4, max_row=3 + len(por_produto))
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)

    ws2.add_chart(chart2, "A9")

    wb.save(caminho_saida)
    print(f"Relatório gerado: {caminho_saida}")


if __name__ == "__main__":
    gerar_relatorio()